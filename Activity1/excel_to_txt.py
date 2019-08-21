from argparse import ArgumentParser
from openpyxl import load_workbook
from os import path

if __name__ == '__main__':
    parser = ArgumentParser(description="Converts excel QS World University Rankings file to a simple txt file")
    parser.add_argument('file', help="QS World University Rankings file to convert")
    args = parser.parse_args()

    file = args.file
    file = path.abspath(file)

    wb = load_workbook(file)
    sheet = wb.active

    new_filename = file.split("/")[-1].replace("xlsx", "txt")
    new_filename = new_filename.replace("-v1.0", "")
    print(new_filename)
    with open(new_filename, "w") as wfile:
        for i in range(6, 206):
            name = sheet.cell(row=i, column=3).value.strip()
            position = str(sheet.cell(row=i, column=1).value) \
                .strip()
            wfile.write(f"{position}, {name}\n")
